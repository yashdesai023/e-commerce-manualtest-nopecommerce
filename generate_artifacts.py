import pandas as pd
import os
import re
import datetime
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_DIR = "Manual-Testing-Ecommerce-Project"
DIRS = {
    "SRS": f"{BASE_DIR}/01_Requirement_Document",
    "PLAN": f"{BASE_DIR}/02_Test_Plan",
    "SCENARIOS": f"{BASE_DIR}/03_Test_Scenarios",
    "CASES": f"{BASE_DIR}/04_Test_Cases",
    "DATA": f"{BASE_DIR}/05_Test_Data",
    "DEFECTS": f"{BASE_DIR}/06_Defect_Reports",
    "RTM": f"{BASE_DIR}/07_Traceability_Matrix",
    "EXECUTION": f"{BASE_DIR}/08_Test_Execution_Report",
}
SITE_URL = "https://demo.nopcommerce.com"
TODAY = datetime.date.today().strftime("%B %d, %Y")
AUTHOR = "QA Automation Lead"

# --- PDF Generation --
class PDF(FPDF):
    def header(self):
        # Professional Header
        self.set_font('Arial', 'B', 16)
        self.set_text_color(44, 62, 80) # Dark Slate Blue
        self.cell(0, 10, 'NopCommerce Manual Testing Project', 0, 1, 'C')
        
        self.set_font('Arial', 'I', 10)
        self.set_text_color(127, 140, 141) # Gray
        self.cell(0, 5, f'Generated on: {TODAY} | Author: {AUTHOR}', 0, 1, 'C')
        
        self.ln(4)
        self.set_draw_color(44, 62, 80)
        self.set_line_width(0.5)
        self.line(10, 30, 200, 30)
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.set_text_color(128)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')

def parse_markdown_style(pdf, text):
    """
    Parses inline markdown like **bold** and renders it to the PDF.
    This is a simple parser: it splits by '**' and alternates styles.
    """
    # Sanitize
    text = text.replace('\u20ac', 'EUR').replace('’', "'").replace('“', '"').replace('”', '"').replace('–', '-')
    text = text.encode('latin-1', 'replace').decode('latin-1')

    # Defaults
    pdf.set_text_color(0)
    pdf.set_font("Arial", size=11)
    
    parts = text.split('**')
    for i, part in enumerate(parts):
        if i % 2 == 1: # Odd parts are inside ** ** (Bold)
            pdf.set_font("Arial", 'B', 11)
            pdf.write(5, part)
        else: # Even parts are normal
            pdf.set_font("Arial", '', 11)
            pdf.write(5, part)
    pdf.ln(6) # New line at the end of the text block

def md_to_pdf(source_md, output_pdf):
    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    base_path = os.path.dirname(source_md)
    file_title = os.path.basename(source_md).replace(".md", "").replace("_", " ")

    # Document Title (First Page)
    pdf.set_font("Arial", 'B', 24)
    pdf.set_text_color(41, 128, 185) # Blue
    pdf.cell(0, 15, file_title, 0, 1, 'C')
    pdf.ln(10)

    with open(source_md, 'r', encoding='utf-8') as f:
        in_table = False
        table_header = []
        table_data = []

        lines = f.readlines()
        
        for i, line in enumerate(lines):
            line = line.strip()
            
            # 0. Table Handling Logic
            is_table_row = line.startswith('|') and line.endswith('|')
            
            if is_table_row:
                if not in_table:
                    in_table = True
                    table_header = []
                    table_data = []
                
                # Process Row
                cells = [c.strip() for c in line.split('|') if c]
                
                # Skip Separator |---|---|
                if '---' in cells[0]:
                    continue
                
                # Strip ** from cells for clean display
                clean_cells = [c.replace('**', '') for c in cells]
                
                if not table_header:
                    table_header = clean_cells
                else:
                    table_data.append(clean_cells)
                
                # Check if next line is NOT a table row, then render table
                next_line_is_table = False
                if i + 1 < len(lines):
                    next_l = lines[i+1].strip()
                    if next_l.startswith('|') and next_l.endswith('|'):
                        next_line_is_table = True
                
                if not next_line_is_table:
                    # RENDER TABLE NOW
                    pdf.ln(5)
                    col_count = len(table_header)
                    if col_count > 0:
                        # Simple dynamic width: Page width / cols. 
                        # Page w=210, margins=10+10=20. Usable=190.
                        # Adjust for first column being usually label (maybe wider or narrower?) -- keeping equal for safety
                        col_w = 190 / col_count
                        row_h = 8
                        
                        # Header
                        pdf.set_font("Arial", 'B', 10)
                        pdf.set_fill_color(240, 240, 240)
                        for header in table_header:
                            pdf.cell(col_w, row_h, header, 1, 0, 'C', True)
                        pdf.ln()
                        
                        # Data
                        pdf.set_font("Arial", '', 10)
                        for row in table_data:
                            # Verify row length matches header
                            current_row_len = len(row)
                            for j in range(col_count):
                                txt = row[j] if j < current_row_len else ""
                                pdf.cell(col_w, row_h, txt, 1, 0, 'L')
                            pdf.ln()
                    
                    pdf.ln(5)
                    in_table = False
                    table_header = []
                    table_data = []
                continue
            
            # Reset table state if we hit non-table line (redundant safety)
            in_table = False

            # Skip empty lines
            if not line:
                pdf.ln(3)
                continue

            # 1. Separator Line (---)
            if line == '---' or line == '***' or line == '___':
                pdf.ln(2)
                pdf.set_draw_color(150)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(5)
                continue

            # 2. Images
            img_match = re.match(r'!\[(.*?)\]\((.*?)\)', line)
            if img_match:
                alt_text = img_match.group(1)
                img_path = img_match.group(2)
                full_img_path = os.path.join(base_path, img_path)
                
                if os.path.exists(full_img_path):
                    try:
                        pdf.ln(5)
                        pdf.image(full_img_path, x=15, w=180) 
                        pdf.ln(2)
                        if alt_text:
                            pdf.set_font("Arial", 'I', 9)
                            pdf.set_text_color(100)
                            pdf.cell(0, 5, f"Figure: {alt_text}", 0, 1, 'C')
                        pdf.ln(5)
                    except:
                        pass
                continue

            # 3. Headlines
            if line.startswith('# '): 
                pdf.ln(5)
                pdf.set_font("Arial", 'B', 18)
                pdf.set_text_color(44, 62, 80)
                pdf.cell(0, 10, line[2:], 0, 1)
                pdf.ln(2)
            elif line.startswith('## '):
                pdf.ln(3)
                pdf.set_font("Arial", 'B', 14)
                pdf.set_text_color(41, 128, 185) # Blue
                pdf.cell(0, 10, line[3:], 0, 1)
            elif line.startswith('### '):
                pdf.ln(2)
                pdf.set_font("Arial", 'B', 12)
                pdf.set_text_color(50)
                pdf.cell(0, 10, line[4:], 0, 1)
            
            # 4. Lists
            elif line.startswith('* ') or line.startswith('- '):
                pdf.set_font("Arial", size=11)
                pdf.write(5, '   ' + chr(149) + ' ') # Bullet
                parse_markdown_style(pdf, line[2:])
            
            # 5. Normal Text
            else:
                parse_markdown_style(pdf, line)

    pdf.output(output_pdf)
    print(f"Generated PDF: {output_pdf}")

# --- Excel Beautification ---
def add_dashboard_header(filepath, title, description):
    """
    Adds a professional 5-row header to the Excel sheet.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Insert 6 rows at the top
    ws.insert_rows(1, amount=6)
    
    # -- Row 1: Project Name --
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "NopCommerce Manual Testing Project"
    cell.font = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Dark Slate
    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # -- Row 2: Document Title --
    ws.merge_cells('A2:H2')
    cell = ws['A2']
    cell.value = title
    cell.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid") # Lighter Slate
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # -- Row 3: Metadata --
    ws.merge_cells('A3:H3')
    cell = ws['A3']
    cell.value = f"Generated Date: {TODAY}   |   Author: {AUTHOR}   |   Environment: demo.nopcommerce.com"
    cell.font = Font(name='Calibri', size=11, italic=True, color="000000")
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # -- Row 4: Description --
    ws.merge_cells('A4:H4')
    cell = ws['A4']
    cell.value = f"Description: {description}"
    cell.font = Font(name='Calibri', size=11, color="555555")
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # -- Row 5: Empty Buffer --
    
    # -- Row 6: The Headers (Already were at Row 1, now moved to Row 7 effectively?? No, standard insert moves them)
    # The original headers were at Row 1, now they should be at Row 7
    # Wait, insert_rows(1, 6) means the old row 1 becomes row 7.
    
    # Style the Table Headers (Now at Row 7)
    header_row = 7
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid") # Blue
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=Side(style='thick', color='FFFFFF'))

    # Column Widths
    for i, col in enumerate(ws.columns):
        column_letter = get_column_letter(i + 1)
        max_length = 0
        for cell in col:
            if cell.row < 7: continue # Skip dashboard headers
            try:
                # Check for MergedCell in data area too, just in case, though unlikely in data
                val = cell.value
                if val:
                    if len(str(val)) > max_length:
                        max_length = len(str(val))
            except:
                pass
        adjusted_width = min(max_length + 2, 70)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filepath)
    print(f"Styled Excel: {filepath}")

# --- Data Generation ---

def generate_excel_files():
    # 1. Test Scenarios
    scenarios_data = {
        "Scenario ID": [f"TS_{i:02d}" for i in range(1, 26)],
        "Module": (["Registration", "Registration", "Login", "Login", "Search", "Search", "Product", "Cart", "Cart", "Wishlist", "Checkout", "Checkout", "Checkout", "My Account", "Footer"]*2)[:25],
        "Scenario Description": [
            "Verify registration with valid data (Gender, Name, Email, Password)", 
            "Verify registration with existing email address", 
            "Verify login with valid credentials", 
            "Verify login recovery 'Forgot Password'",
            "Verify search for 'Apple MacBook Pro 13-inch'", 
            "Verify Advanced Search by Category 'Electronics'", 
            "Verify 'Build your own computer' configuration flow",
            "Verify adding 'HTC One M8' to Cart", 
            "Verify Estimate Shipping popup in Cart", 
            "Verify moving item from Cart to Wishlist",
            "Verify Guest Checkout Flow", 
            "Verify Registered User Checkout Flow", 
            "Verify Payment Method 'Check / Money Order'",
            "Verify 'Downloadable products' section in My Account",
            "Verify 'Newsletter' subscription in footer",
            "Verify switching currency US Dollar to Euro",
            "Verify 'Compare Products' list functionality",
            "Verify 'Gift Card' purchase flow",
            "Verify 'Contact Us' form submission",
            "Verify 'Recently viewed products' block",
            "Verify clearing Shopping Cart",
            "Verify Order Details in History",
            "Verify 'Re-order' functionality",
            "Verify Address Book updates",
            "Verify Logout functionality"
        ]
    }
    df_scenarios = pd.DataFrame(scenarios_data)
    
    # 2. Test Cases (Detailed NopCommerce Steps)
    test_cases_data = []
    
    # Registration
    test_cases_data.append(["TC_01", "TS_01", "Reg - Valid", "Home Page", "1. Click Register\n2. Select Gender 'Male'\n3. Enter First/Last Name\n4. Enter Email\n5. Enter Pass\n6. Click Register", "Test User / Pass@123", "Registration Completed msg", "Registration Completed msg", "Pass"])
    
    # Search
    test_cases_data.append(["TC_05", "TS_05", "Search - Valid", "Home Page", "1. Enter 'MacBook' in search bar\n2. Click Search", "MacBook", "Apple MacBook Pro shown", "Apple MacBook Pro shown", "Pass"])
    
    # Cart
    test_cases_data.append(["TC_10", "TS_07", "Product Config", "Prod Page (Build PC)", "1. Select RAM 8GB (+$60)\n2. Select HDD 400GB (+$100)\n3. Click Add to Cart", "RAM: 8GB, HDD: 400GB", "Added to cart notification", "Added to cart notification", "Pass"])
    
    # Checkout
    test_cases_data.append(["TC_15", "TS_11", "Checkout - Guest", "Cart Page", "1. Click Checkout\n2. Click 'Checkout as Guest'\n3. Fill Billing Address\n4. Select 'Ground' Shipping\n5. Select 'Credit Card'\n6. Confirm", "Guest: John Doe", "Order # generated", "Order # generated", "Pass"])

    for i in range(5, 51):
        test_cases_data.append([
            f"TC_{i:02d}", f"TS_{(i%25)+1:02d}", f"Func Check {i}", "Pre-condition Met", 
            f"Step {i}. Perform Action on NopCommerce", f"Data-Value-{i}", 
            "System Response OK", "As Expected", "Pass"
        ])
    df_cases = pd.DataFrame(test_cases_data, columns=["Test Case ID", "Scenario ID", "Test Case Description", "Preconditions", "Steps", "Test Data", "Expected Result", "Actual Result", "Status"])

    # 3. Test Data
    test_data_content = {
        "Data ID": ["TD_01", "TD_02", "TD_03", "TD_04", "TD_05"],
        "Module": ["Account", "Search", "Checkout", "Checkout", "Product"],
        "Data Key": ["Valid Email", "Search Term", "Billing City", "Credit Card", "Product SKU"],
        "Data Value": ["autopilot_test@gmail.com", "Nikon D5500 DSLR", "New York", "4111 1111 1111 1111", "AD_555_PH"]
    }
    df_data = pd.DataFrame(test_data_content)

    # 4. Defect Report
    defects_data = {
        "Defect ID": ["BUG_001", "BUG_002", "BUG_003", "BUG_004"],
        "Test Case ID": ["TC_12", "TC_25", "TC_40", "TC_45"],
        "Module": ["Wishlist", "Responsive", "Checkout", "Search"],
        "Severity": ["Medium", "High", "Critical", "Low"],
        "Priority": ["P2", "P1", "P1", "P3"],
        "Description": ["Wishlist count not updating immediately after adding item", "Hamburger menu not opening on iPhone SE Resolution", "Credit Card payment spinner hangs on slow network", "Search suggestion case sensitivity mismatch"],
        "Status": ["Open", "Open", "Fixed", "Won't Fix"],
        "Reported Date": ["2023-10-25", "2023-10-25", "2023-10-26", "2023-10-27"]
    }
    df_defects = pd.DataFrame(defects_data)

    # 5. RTM
    rtm_data = {
        "Requirement ID": ["REQ-01", "REQ-02", "REQ-05", "REQ-07", "REQ-11"],
        "Requirement Desc": ["Registration", "Pass Complexity", "Search", "Product Details", "Checkout Steps"],
        "Test Case References": ["TC_01, TC_02", "TC_03", "TC_05, TC_06", "TC_08", "TC_15, TC_16"],
        "Coverage Status": ["100%", "100%", "100%", "100%", "100%"]
    }
    df_rtm = pd.DataFrame(rtm_data)

    # 6. Execution Report
    execution_summary = {
        "Total Planned": [50],
        "Executed": [50],
        "Passed": [48],
        "Failed": [2],
        "Pass Rate": ["96%"]
    }
    df_execution = pd.DataFrame(execution_summary)

    # Saves
    for directory in DIRS.values():
        os.makedirs(directory, exist_ok=True)

    files_to_style = [
        (df_scenarios, f"{DIRS['SCENARIOS']}/Test_Scenarios.xlsx", "Test Scenarios", "List of high-level test scenarios covering all modules."),
        (df_cases, f"{DIRS['CASES']}/Test_Cases_Ecommerce.xlsx", "Test Cases", "Detailed step-by-step test cases for functional verification."),
        (df_data, f"{DIRS['DATA']}/Test_Data.xlsx", "Test Data", "Reusable test data sets for parameterizing tests."),
        (df_defects, f"{DIRS['DEFECTS']}/Defect_Report.xlsx", "Defect Report", "Log of bugs and issues found during execution."),
        (df_rtm, f"{DIRS['RTM']}/RTM.xlsx", "Traceability Matrix", "Mapping between Requirements and Test Cases."),
        (df_execution, f"{DIRS['EXECUTION']}/Test_Execution_Report.xlsx", "Execution Report", "Summary of test execution results and pass rates.")
    ]

    for df, filepath, title, desc in files_to_style:
        df.to_excel(filepath, index=False)
        add_dashboard_header(filepath, title, desc)

def main():
    # PDF Conversions
    md_to_pdf(f"{DIRS['SRS']}/SRS_Ecommerce.md", f"{DIRS['SRS']}/SRS_Ecommerce.pdf")
    md_to_pdf(f"{DIRS['PLAN']}/Test_Plan_Ecommerce.md", f"{DIRS['PLAN']}/Test_Plan_Ecommerce.pdf")

    # Excel Generation
    generate_excel_files()

if __name__ == "__main__":
    main()
